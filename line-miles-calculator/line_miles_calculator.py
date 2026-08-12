"""
Line Miles Calculation & Delivery-Time Estimation, by Feeder
------------------------------------------------------------
Aggregates line-length attributes across a spatial network, grouped by
feeder, and produces a per-feeder Excel breakdown — including an
estimated delivery time based on which validation method each feature
requires.

Category model: reference-source availability, not an arbitrary
difficulty tier
--------------------------------------------------------------------
Some features can be cross-validated against recent street-level
imagery in addition to satellite/aerial imagery; others have no
street-level reference (or one too old to trust) and must be
interpreted from imagery alone. Cross-validation takes longer per
line-mile but produces higher-confidence output — this is a genuine,
explainable driver of delivery time, not an arbitrary label.

Two ways to classify each feature are supported:

1. `category_field` - the dataset already carries a pre-assigned
   category value (e.g. an existing label-flag attribute). Used as-is.
2. `date_field` - the dataset carries a capture-date attribute for the
   street-level reference. The tool derives the category itself by
   comparing that date against `recency_threshold_years` (relative to
   today, or a supplied reference date). "Recency" is a moving target,
   so this threshold is a parameter, not a hardcoded constant.

Honesty about missing data
---------------------------
If neither a usable category value nor a parseable date is available
for a feature, it is classified as "Unknown" — its length still counts
toward the feeder's total line miles, but it is excluded from the
time estimate and reported separately as a data-quality gap. The tool
never guesses a category to fill in missing data.

Honesty about the time estimate itself
----------------------------------------
The per-category time-per-mile rates are a planning-level heuristic for
resourcing and scheduling, not a precise prediction. The defaults below
are illustrative starting points — calibrate them against your own
historical delivery data before relying on them for real scheduling.

Requirements
------------
- Standalone mode: pandas, openpyxl (for Excel export)
- QGIS mode: run inside QGIS (uses qgis.core, bundled)
"""

from collections import defaultdict
from datetime import datetime, date

import pandas as pd


# ---------------------------------------------------------------------------
# Defaults — illustrative only, calibrate from your own historical data
# ---------------------------------------------------------------------------

DEFAULT_RECENCY_THRESHOLD_YEARS = 2
DEFAULT_CATEGORY_LABELS = ("Cross-Validated", "Imagery-Only")
DEFAULT_TIME_ESTIMATES_MIN_PER_MILE = {
    "Cross-Validated": 8.0,   # slower — validated against two independent sources
    "Imagery-Only": 2.5,      # faster — single-source interpretation
}
UNKNOWN_LABEL = "Unknown"


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

def classify_by_recency(capture_date, reference_date=None,
                         recency_threshold_years=DEFAULT_RECENCY_THRESHOLD_YEARS,
                         labels=DEFAULT_CATEGORY_LABELS):
    """
    Classify a feature as recently cross-validated or imagery-only,
    based on a street-level-reference capture date.

    Parameters
    ----------
    capture_date : datetime, date, str, or None
        The street-level reference's capture date. None/unparseable
        returns UNKNOWN_LABEL.
    reference_date : date, optional
        "Today", for testing/reproducibility. Defaults to the real
        current date.
    recency_threshold_years : float
        How many years back still counts as "recent". This is a
        parameter because what counts as "latest" shifts every year.
    labels : (recent_label, stale_label)

    Returns
    -------
    str
        labels[0] if within the recency threshold, labels[1] if not,
        UNKNOWN_LABEL if capture_date is missing/unparseable.
    """
    if capture_date is None or (isinstance(capture_date, float) and pd.isna(capture_date)):
        return UNKNOWN_LABEL

    if isinstance(capture_date, str):
        try:
            capture_date = pd.to_datetime(capture_date).date()
        except (ValueError, TypeError):
            return UNKNOWN_LABEL
    elif isinstance(capture_date, datetime):
        capture_date = capture_date.date()
    elif not isinstance(capture_date, date):
        return UNKNOWN_LABEL

    reference_date = reference_date or datetime.now().date()
    age_years = (reference_date - capture_date).days / 365.25

    return labels[0] if age_years <= recency_threshold_years else labels[1]


# ---------------------------------------------------------------------------
# MODE 1: Standalone — works on a pandas DataFrame (from CSV, Excel, or
# an attribute table exported from any GIS)
# ---------------------------------------------------------------------------

def calculate_line_miles_by_feeder(
    df,
    length_field="length_lms",
    feeder_field="feeder",
    customer_field=None,
    category_field=None,
    date_field=None,
    recency_threshold_years=DEFAULT_RECENCY_THRESHOLD_YEARS,
    category_labels=DEFAULT_CATEGORY_LABELS,
    time_estimates_min_per_mile=None,
    reference_date=None,
):
    """
    Aggregate line miles by feeder and validation category.

    Parameters
    ----------
    df : pandas.DataFrame
        One row per spatial feature, with at least a length and a
        feeder-identifier column.
    length_field, feeder_field : str
        Column names for line length and feeder ID.
    customer_field : str, optional
        Column name for a customer/client identifier, included in the
        output if present.
    category_field : str, optional
        Column already holding a pre-assigned category label. Takes
        precedence over date_field if both are given.
    date_field : str, optional
        Column holding the street-level-reference capture date. Used to
        derive the category via `classify_by_recency` when
        category_field isn't provided (or is null for a given row).
    recency_threshold_years : float
        Passed to `classify_by_recency`.
    category_labels : tuple
        (recent_label, stale_label) passed to `classify_by_recency`.
    time_estimates_min_per_mile : dict, optional
        {category_label: minutes_per_mile}. Defaults to
        DEFAULT_TIME_ESTIMATES_MIN_PER_MILE. Categories not present in
        this dict (including "Unknown") are excluded from time
        estimates but still counted in total line miles.
    reference_date : date, optional
        "Today", for testing/reproducibility.

    Returns
    -------
    dict
        {feeder_name: {
            'customer': str,
            'lengths': {category: total_length, ...},
            'counts': {category: feature_count, ...},
            'estimated_minutes': {category: minutes, ...},
            'unknown_length': float,
            'unknown_count': int,
        }}
    """
    if category_field is None and date_field is None:
        raise ValueError("Provide at least one of category_field or date_field")

    time_estimates = time_estimates_min_per_mile or DEFAULT_TIME_ESTIMATES_MIN_PER_MILE

    results = defaultdict(lambda: {
        "customer": "", "lengths": defaultdict(float), "counts": defaultdict(int),
    })

    skipped = 0
    for _, row in df.iterrows():
        length_val = row.get(length_field)
        feeder = row.get(feeder_field)

        if pd.isna(length_val) or pd.isna(feeder):
            skipped += 1
            continue
        try:
            length_val = float(length_val)
        except (ValueError, TypeError):
            skipped += 1
            continue

        feeder = str(feeder).strip()

        category = None
        if category_field is not None:
            raw = row.get(category_field)
            if raw is not None and not (isinstance(raw, float) and pd.isna(raw)):
                category = str(raw).strip()
        if category is None and date_field is not None:
            category = classify_by_recency(
                row.get(date_field), reference_date=reference_date,
                recency_threshold_years=recency_threshold_years, labels=category_labels,
            )
        if category is None:
            category = UNKNOWN_LABEL

        if customer_field is not None:
            cust = row.get(customer_field)
            if cust is not None and not (isinstance(cust, float) and pd.isna(cust)):
                results[feeder]["customer"] = str(cust).strip()

        results[feeder]["lengths"][category] += length_val
        results[feeder]["counts"][category] += 1

    if skipped:
        print(f"Skipped {skipped} row(s) with missing/invalid length or feeder value.")

    # Compute estimated time per category, and surface data-quality gaps
    for feeder, data in results.items():
        data["estimated_minutes"] = {}
        for category, length in data["lengths"].items():
            if category in time_estimates:
                data["estimated_minutes"][category] = length * time_estimates[category]
        data["unknown_length"] = data["lengths"].get(UNKNOWN_LABEL, 0.0)
        data["unknown_count"] = data["counts"].get(UNKNOWN_LABEL, 0)

    return dict(results)


def export_to_excel(results, output_path, time_estimates_min_per_mile=None):
    """
    Export per-feeder results to a formatted Excel workbook: a summary
    sheet, a data-quality sheet flagging Unknown-category features, and
    a metadata sheet.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    time_estimates = time_estimates_min_per_mile or DEFAULT_TIME_ESTIMATES_MIN_PER_MILE
    all_categories = sorted({
        cat for data in results.values() for cat in data["lengths"]
        if cat != UNKNOWN_LABEL
    })

    wb = Workbook()
    ws = wb.active
    ws.title = "Feeder Summary"

    headers = (["Customer", "Feeder", "Span Count", "Total Line Miles"]
               + [f"{c} (miles)" for c in all_categories]
               + [f"{c} (est. hrs)" for c in all_categories]
               + ["Unknown (miles)", "Unknown Count"])
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    grand_totals = defaultdict(float)
    for feeder in sorted(results.keys()):
        data = results[feeder]
        total_miles = sum(data["lengths"].values())
        total_count = sum(data["counts"].values())
        row = [data["customer"], feeder, total_count, round(total_miles, 4)]
        for c in all_categories:
            row.append(round(data["lengths"].get(c, 0.0), 4))
        for c in all_categories:
            hrs = data["estimated_minutes"].get(c, 0.0) / 60.0
            row.append(round(hrs, 2))
        row.append(round(data["unknown_length"], 4))
        row.append(data["unknown_count"])
        ws.append(row)

        grand_totals["miles"] += total_miles
        grand_totals["unknown_miles"] += data["unknown_length"]
        grand_totals["unknown_count"] += data["unknown_count"]

    for col in ws.columns:
        width = max(len(str(c.value)) for c in col if c.value is not None) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 30)

    # Data-quality sheet
    dq = wb.create_sheet("Data Quality")
    dq.append(["Feeder", "Unknown-category Span Count", "Unknown-category Line Miles"])
    for cell in dq[1]:
        cell.font = Font(bold=True)
    any_unknown = False
    for feeder in sorted(results.keys()):
        data = results[feeder]
        if data["unknown_count"] > 0:
            any_unknown = True
            dq.append([feeder, data["unknown_count"], round(data["unknown_length"], 4)])
    if not any_unknown:
        dq.append(["No features with missing/unusable classification data."])

    # Metadata sheet
    meta = wb.create_sheet("Metadata")
    meta.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    meta.append(["Feeders:", len(results)])
    meta.append(["Total line miles:", round(grand_totals["miles"], 4)])
    meta.append(["Unknown-category line miles:", round(grand_totals["unknown_miles"], 4)])
    meta.append([""])
    meta.append(["Time-estimate rates (minutes/mile) — illustrative, calibrate from your own data:"])
    for cat, rate in time_estimates.items():
        meta.append([cat, rate])

    wb.save(output_path)
    print(f"Exported: {output_path}")
    return output_path


def run_qgis_batch(
    length_field="length_lms",
    feeder_field=None,
    customer_field=None,
    category_field=None,
    date_field=None,
    recency_threshold_years=DEFAULT_RECENCY_THRESHOLD_YEARS,
    output_path=None,
):
    """
    Run the same feeder aggregation against the currently active QGIS
    layer's attribute table. Auto-detects common feeder/customer field
    name variants if not specified, mirroring the standalone mode's
    output exactly (same underlying function, just a different data
    source).
    """
    from qgis.core import QgsProject, QgsMapLayer
    from qgis.utils import iface

    layer = iface.activeLayer()
    if not layer:
        print("No active layer selected.")
        return None
    if layer.type() != QgsMapLayer.VectorLayer:
        print(f"'{layer.name()}' is a {layer.typeName()} layer — this tool needs a vector layer.")
        return None

    field_names = [f.name() for f in layer.fields()]

    if feeder_field is None:
        feeder_field = next((f for f in ["feeder", "feeder_name", "feeder_id", "name"] if f in field_names), None)
        if feeder_field is None:
            print(f"Could not auto-detect a feeder field. Available fields: {field_names}")
            return None

    if customer_field is None:
        customer_field = next(
            (f for f in ["customer", "customer_name", "customer_id", "cust_name", "client", "client_name"]
             if f in field_names), None,
        )

    print(f"Layer: {layer.name()} | Feeder field: {feeder_field} | Customer field: {customer_field or 'none'}")

    rows = []
    for feature in layer.getFeatures():
        row = {f: feature[f] for f in field_names if f in
               {length_field, feeder_field, customer_field, category_field, date_field}}
        rows.append(row)

    df = pd.DataFrame(rows)
    results = calculate_line_miles_by_feeder(
        df, length_field=length_field, feeder_field=feeder_field, customer_field=customer_field,
        category_field=category_field, date_field=date_field, recency_threshold_years=recency_threshold_years,
    )

    if output_path:
        export_to_excel(results, output_path)

    return results


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Calculate line miles by feeder with recency-based delivery-time estimation.")
    parser.add_argument("input_csv", help="CSV/Excel attribute table with length, feeder, and date/category columns")
    parser.add_argument("output_xlsx", help="Output Excel path")
    parser.add_argument("--length-field", default="length_lms")
    parser.add_argument("--feeder-field", default="feeder")
    parser.add_argument("--customer-field", default=None)
    parser.add_argument("--category-field", default=None)
    parser.add_argument("--date-field", default=None)
    parser.add_argument("--recency-years", type=float, default=DEFAULT_RECENCY_THRESHOLD_YEARS)
    args = parser.parse_args()

    data = pd.read_csv(args.input_csv) if args.input_csv.endswith(".csv") else pd.read_excel(args.input_csv)
    results = calculate_line_miles_by_feeder(
        data, length_field=args.length_field, feeder_field=args.feeder_field,
        customer_field=args.customer_field, category_field=args.category_field,
        date_field=args.date_field, recency_threshold_years=args.recency_years,
    )
    export_to_excel(results, args.output_xlsx)
