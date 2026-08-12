"""
Polygon Delta QC Engine
------------------------
Compares two folders of polygon shapefiles — a "reference" (e.g. model
output) dataset and a "comparison" (e.g. QC/vendor-corrected) dataset —
organized one subfolder per feeder/tile, and reports how many unique
features were added or removed between them, per feeder and per
feature ID.

Originally built for tree-health polygon QC (comparing an ML model's
tree-health output against a human-QC-corrected version to measure how
much correction was needed), but the underlying comparison technique is
generic: any workflow that needs to reconcile two versions of a
categorized polygon dataset — crop-boundary corrections, building
footprint QC, land-parcel review — fits the same pattern.

Two interfaces are provided:

1. Headless (`compare_polygon_datasets`) - no GUI dependency, callable
   from any script or pipeline. This is what the GUI calls internally.
2. GUI (`PolygonDeltaQCEngine`, PyQt5) - a desktop app wrapping the
   headless function in a background thread, with folder browsers, a
   live log, and a progress bar, for interactive use.

Counting methodology
---------------------
A feature is identified by the combination of an ID column and an area
column (defaults: "treeid", "treearea" - configurable to match your own
schema). Two rows with the same ID and area are treated as one feature
(duplicates collapse); the same ID with a different area is treated as
a separate feature. Rows where every attribute is null/empty are still
counted individually (each null polygon is a real feature that needs
review, not noise to discard).

"Deletion" = reference count - comparison count, i.e. how many unique
features were present in the reference set but not confirmed in the
comparison set. This is a count-based delta, not a duplicate-of-truth
answer key — it measures correction volume, not correctness (assuming the
comparison set is the corrected one).

Requirements
------------
- Headless core: pandas, openpyxl (only geopandas/pyshp/dbfread are
  *optional* fallbacks for the shapefile reader — the primary path
  parses the DBF binary format directly, no extra dependency required)
- GUI: PyQt5
"""

import os
import struct
import pandas as pd
from collections import defaultdict


DEFAULT_ID_COLUMN = "treeid"
DEFAULT_AREA_COLUMN = "treearea"


class ShapefileReader:
    """Robust shapefile attribute-table reader with multiple fallback methods."""

    @staticmethod
    def read_dbf_raw(dbf_path):
        """Read a DBF file's attribute table via direct binary parsing —
        no external dependency required. Most reliable method; tried first."""
        try:
            with open(dbf_path, "rb") as f:
                f.seek(4)
                num_records = struct.unpack("<I", f.read(4))[0]
                f.seek(8)
                header_length = struct.unpack("<H", f.read(2))[0]
                record_length = struct.unpack("<H", f.read(2))[0]

                f.seek(32)
                fields = []
                while True:
                    field_info = f.read(32)
                    if field_info[0] == 0x0D:
                        break
                    field_name = field_info[:11].decode("utf-8").strip("\x00")
                    field_type = chr(field_info[11])
                    field_length = field_info[16]
                    fields.append({"name": field_name, "type": field_type, "length": field_length})

                f.seek(header_length)
                records = []
                for _ in range(num_records):
                    deletion_flag = f.read(1)
                    if deletion_flag == b"*":
                        f.read(record_length - 1)
                        continue

                    record = {}
                    for field in fields:
                        value = f.read(field["length"]).decode("utf-8", errors="ignore").strip()
                        if field["type"] == "N":
                            try:
                                value = (float(value) if "." in value else int(value)) if value else None
                            except ValueError:
                                value = None
                        elif field["type"] == "L":
                            value = value.upper() in ("T", "Y")
                        elif field["type"] == "D" and value:
                            try:
                                value = pd.to_datetime(value, format="%Y%m%d")
                            except ValueError:
                                pass
                        record[field["name"]] = value
                    records.append(record)

                return pd.DataFrame(records)
        except Exception as e:
            raise Exception(f"Failed to read DBF: {str(e)}")

    @staticmethod
    def read_shapefile(shapefile_path):
        """Read a shapefile's attribute table, trying the dependency-free
        raw DBF parser first, then falling back to geopandas / pyshp /
        dbfread if that fails."""
        dbf_path = shapefile_path.replace(".shp", ".dbf")
        if not os.path.exists(dbf_path):
            raise Exception(f"DBF file not found: {dbf_path}")

        try:
            return ShapefileReader.read_dbf_raw(dbf_path)
        except Exception:
            pass
        try:
            import geopandas as gpd
            gdf = gpd.read_file(shapefile_path)
            return pd.DataFrame(gdf.drop(columns="geometry"))
        except Exception:
            pass
        try:
            import shapefile
            sf = shapefile.Reader(shapefile_path)
            records = [r.as_dict() for r in sf.records()]
            return pd.DataFrame(records)
        except Exception:
            pass
        try:
            from dbfread import DBF
            table = DBF(dbf_path, encoding="utf-8", ignore_missing_memofile=True)
            return pd.DataFrame(iter(table))
        except Exception:
            pass

        raise Exception("All shapefile reading methods failed. The file may be corrupted.")


# ---------------------------------------------------------------------------
# Headless core — no GUI dependency, testable in any Python environment
# ---------------------------------------------------------------------------

def _is_polygon_shapefile(shapefile_path):
    try:
        with open(shapefile_path, "rb") as f:
            f.seek(32)
            geom_type = struct.unpack("<I", f.read(4))[0]
            return geom_type in [5, 15, 25]  # Polygon shape types
    except Exception:
        return True


def _is_bbox_shapefile(filename):
    indicators = ["bbox", "bounding", "boundary", "extent", "clip", "mask", "border"]
    return any(ind in filename.lower() for ind in indicators)


def _has_required_columns(shapefile_path, required_columns):
    try:
        df = ShapefileReader.read_shapefile(shapefile_path)
        if df is None or len(df) == 0:
            return False
        cols_lower = [c.lower() for c in df.columns.tolist()]
        return all(c.lower() in cols_lower for c in required_columns)
    except Exception:
        return False


def find_polygon_shapefiles(base_folder, id_column, area_column, log=print):
    """Recursively find one polygon shapefile per feeder/tile subfolder,
    skipping bbox files and requiring the ID/area columns to be present."""
    required = [id_column, area_column]
    found = {}

    for root, _dirs, files in os.walk(base_folder):
        shp_files = [f for f in files if f.lower().endswith(".shp")]
        feeder_name = os.path.basename(base_folder) if root == base_folder else os.path.basename(root)

        if feeder_name in found:
            continue

        candidates = []
        for shp_file in shp_files:
            shp_path = os.path.join(root, shp_file)
            if _is_bbox_shapefile(shp_file):
                continue
            if not _is_polygon_shapefile(shp_path):
                continue
            if _has_required_columns(shp_path, required):
                candidates.append((shp_path, shp_file))

        if candidates:
            non_line = [c for c in candidates if "line" not in c[1].lower()]
            selected_path, selected_file = (non_line or candidates)[0]
            found[feeder_name] = selected_path
            log(f"  Feeder '{feeder_name}': {os.path.relpath(selected_path, base_folder)}")

    return found


def count_unique_features(df, id_column, area_column):
    """
    Count unique (id, area) combinations. Rows where every attribute is
    null/empty are still counted, individually, as separate features
    (a null polygon is a real feature needing review, not noise).

    Returns
    -------
    (actual_row_count, unique_feature_count, set of (id, area) tuples)
    """
    if df is None or len(df) == 0:
        return 0, 0, set()

    combos = set()
    null_counter = 0

    for _, row in df.iterrows():
        fid = row.get(id_column)
        area = row.get(area_column)
        all_null = all(pd.isna(v) or str(v).strip() == "" for v in row.values)

        if all_null:
            null_counter += 1
            combos.add(("__NULL_POLYGON__", f"__NULL_AREA_{null_counter}__"))
        elif pd.notna(fid) and pd.notna(area):
            combos.add((fid, area))

    return len(df), len(combos), combos


def compare_polygon_datasets(
    reference_folder,
    comparison_folder,
    output_path,
    id_column=DEFAULT_ID_COLUMN,
    area_column=DEFAULT_AREA_COLUMN,
    reference_label="Reference",
    comparison_label="Comparison",
    log=print,
    progress_callback=None,
):
    """
    Compare two folders of per-feeder polygon shapefiles and write a
    formatted Excel delta report.

    Parameters
    ----------
    reference_folder, comparison_folder : str
        Folders containing one polygon shapefile per feeder/tile
        subfolder (e.g. an ML model's output vs. a QC-corrected version).
    output_path : str
        Path to write the Excel report to.
    id_column, area_column : str
        Attribute names identifying a unique feature. Defaults match
        the tool's original tree-health use case ("treeid", "treearea")
        but any two columns that together identify a feature will work.
    reference_label, comparison_label : str
        Column-header labels in the report (e.g. "Model"/"QC",
        "Draft"/"Final", or your own terms).
    log : callable
        Called with progress-log strings (default: print).
    progress_callback : callable, optional
        Called with (current_index, total) as each feeder is processed
        — useful for driving a GUI progress bar.

    Returns
    -------
    str
        Path to the written Excel report, or None if nothing was processed.
    """
    log("Locating polygon shapefiles...")
    ref_shapefiles = find_polygon_shapefiles(reference_folder, id_column, area_column, log)
    comp_shapefiles = find_polygon_shapefiles(comparison_folder, id_column, area_column, log)

    if not ref_shapefiles:
        log(f"No valid polygon shapefiles found in {reference_folder}")
        return None
    if not comp_shapefiles:
        log(f"No valid polygon shapefiles found in {comparison_folder}")
        return None

    all_feeders = sorted(set(ref_shapefiles) | set(comp_shapefiles))
    log(f"Feeders to process: {len(all_feeders)}")

    ref_actual, ref_unique = {}, {}
    comp_actual, comp_unique = {}, {}
    ref_by_id = defaultdict(lambda: defaultdict(int))
    comp_by_id = defaultdict(lambda: defaultdict(int))
    processed = []

    for idx, feeder in enumerate(all_feeders, 1):
        if progress_callback:
            progress_callback(idx, len(all_feeders))
        did_process = False

        for shapefiles, actual_d, unique_d, by_id_d, label in (
            (ref_shapefiles, ref_actual, ref_unique, ref_by_id, reference_label),
            (comp_shapefiles, comp_actual, comp_unique, comp_by_id, comparison_label),
        ):
            if feeder not in shapefiles:
                continue
            df = ShapefileReader.read_shapefile(shapefiles[feeder])
            if df is None or len(df) == 0:
                continue
            df.columns = [c.lower() for c in df.columns]
            if id_column not in df.columns or area_column not in df.columns:
                log(f"  [{feeder}] {label}: missing required columns, skipping")
                continue

            actual, unique, combos = count_unique_features(df, id_column, area_column)
            actual_d[feeder], unique_d[feeder] = actual, unique
            for fid, area in combos:
                by_id_d[feeder][fid] += 1
            did_process = True

        if did_process:
            processed.append(feeder)

    if not processed:
        log("No feeders were successfully processed.")
        return None

    _write_report(
        output_path, processed, ref_actual, ref_unique, comp_actual, comp_unique,
        ref_by_id, comp_by_id, reference_label, comparison_label,
    )
    log(f"Report written: {output_path}")
    return output_path


def _write_report(output_path, feeders, ref_actual, ref_unique, comp_actual, comp_unique,
                   ref_by_id, comp_by_id, reference_label, comparison_label):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    border = Border(*(Side(style="thin", color="000000"),) * 4)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_rows = []
        for feeder in feeders:
            r_count = ref_unique.get(feeder, 0)
            c_count = comp_unique.get(feeder, 0)
            deletion = r_count - c_count
            diff_pct = round((deletion / r_count) * 100) if r_count > 0 else 0
            summary_rows.append({
                "Feeder": feeder,
                f"{reference_label} Output": ref_actual.get(feeder, 0),
                f"{comparison_label} Output": comp_actual.get(feeder, 0),
                f"{reference_label} Count": r_count,
                f"{comparison_label} Count": c_count,
                "Deletion": deletion,
                "Difference %": diff_pct,
            })

        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]
        _style_sheet(ws, summary_df, header_fill, header_font, header_align, border, pct_col=7)

        for feeder in feeders:
            all_ids = sorted(set(ref_by_id[feeder]) | set(comp_by_id[feeder]))
            rows = [{
                "ID": fid,
                f"{reference_label} Count": ref_by_id[feeder].get(fid, 0),
                f"{comparison_label} Count": comp_by_id[feeder].get(fid, 0),
                "Deletion": ref_by_id[feeder].get(fid, 0) - comp_by_id[feeder].get(fid, 0),
            } for fid in all_ids]

            df = pd.DataFrame(rows)
            sheet_name = str(feeder)[:31]
            for ch in ['/', '\\', '*', '?', '[', ']', ':']:
                sheet_name = sheet_name.replace(ch, "_")
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer.sheets[sheet_name], df, header_fill, header_font, header_align, border)


def _style_sheet(ws, df, header_fill, header_font, header_align, border, pct_col=None):
    from openpyxl.styles import Alignment

    for col_num in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, header_align, border

    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if pct_col and col_idx == pct_col:
                cell.number_format = '0"%"'
                cell.alignment = Alignment(horizontal="center")

    for column in ws.columns:
        max_len = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max_len + 2, 50)
